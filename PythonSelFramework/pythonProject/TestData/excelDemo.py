import openpyxl
book=openpyxl.load_workbook("D:\\Python Udemy\\PythonSelFramework\\demoxlpython.xlsx")
sheet=book.active
cell=sheet.cell(row=1,column=2)
print(cell.value)
sheet.cell(row=2,column=2).value="Pratik"
print(sheet.cell(row=2,column=2).value)
print(sheet.max_row)
print(sheet.max_column)
print(sheet['A3'].value)
Dict={}
for i in range(1,sheet.max_row+1):
    for j in range(1,sheet.max_column+1):
        print(sheet.cell(row=i,column=j).value)

print("to search only particular row data")
for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value=="Testcase1":
        for j in range(1,sheet.max_column+1):
            print(sheet.cell(row=i,column=j).value)

print("store data in a dic")

for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value=="Testcase1":
        for j in range(2,sheet.max_column+1):
            Dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value

print(Dict)

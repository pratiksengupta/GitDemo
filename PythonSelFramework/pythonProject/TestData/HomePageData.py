import openpyxl


class HomePageData:
    test_HomePage_Data=[{"firstname":"Pratik","email":"hello@gmail.com","password":"123456","gender":1,"text":"test"},{"firstname":"Moumita","email":"hello2@gmail.com","password":"654321","gender":0,"text":"test"}]

    @staticmethod
    def getData(test_case_name):
        book = openpyxl.load_workbook("D:\\Python Udemy\\PythonSelFramework\\demoxlpython.xlsx")
        sheet = book.active
        Dict = {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]


